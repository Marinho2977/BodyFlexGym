from flask import Flask, render_template, request, redirect, session, flash, make_response
from werkzeug.security import generate_password_hash, check_password_hash
import mysql.connector
from datetime import date, timedelta, datetime
import os
import secrets
import smtplib
import ssl
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from io import BytesIO
import re
import math
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
from reportlab.lib.colors import HexColor

app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", "clave_segura_de_kevin_123")

# ─────────────────────────────────────────────
# CSRF PROTECTION
# ─────────────────────────────────────────────

def generar_csrf_token():
    if "_csrf_token" not in session:
        session["_csrf_token"] = secrets.token_hex(32)
    return session["_csrf_token"]

app.jinja_env.globals["csrf_token"] = generar_csrf_token

@app.before_request
def verificar_csrf():
    if request.method == "POST":
        token_sesion = session.get("_csrf_token")
        token_form   = request.form.get("csrf_token")
        if not token_sesion or token_sesion != token_form:
            flash("Sesión inválida. Intenta de nuevo.", "error")
            return redirect(request.referrer or "/login")

@app.after_request
def sin_cache(response):
    response.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
    response.headers["Pragma"] = "no-cache"
    response.headers["Expires"] = "0"
    return response

# --- CONFIGURACIÓN DE BASE DE DATOS ---
def conectar_db():
    return mysql.connector.connect(
        host=os.environ.get('MYSQLHOST'),
        user=os.environ.get('MYSQLUSER'),
        password=os.environ.get('MYSQLPASSWORD'),
        database=os.environ.get('MYSQLDATABASE'),
        port=int(os.environ.get('MYSQLPORT', 3306))
    )

PRECIO_MENSUAL = 225.00
MESES_NOMBRES = ["Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
                 "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]
MESES_POR_NOMBRE = {nombre.lower(): i + 1 for i, nombre in enumerate(MESES_NOMBRES)}

def calcular_monto_pago(meses):
    cant_6 = meses // 6
    resto_6 = meses % 6
    cant_3 = resto_6 // 3
    resto_3 = resto_6 % 3
    return float((cant_6 * 1100) + (cant_3 * 600) + (resto_3 * 225))

def calcular_fecha_vencimiento_dia_3(anio, mes):
    siguiente_anio = anio + (mes // 12)
    siguiente_mes = (mes % 12) + 1
    return date(siguiente_anio, siguiente_mes, 3)


# --- CONFIGURACIÓN DE CORREO ---
GMAIL_USER     = os.environ.get("GMAIL_USER")
GMAIL_PASSWORD = os.environ.get("GMAIL_PASSWORD")

# ─────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────

def registrar_log(tipo, detalle, afectado_id=None, afectado_nombre=None):
    actor_cui    = session.get("usuario_id")
    actor_nombre = session.get("nombre", "Sistema")
    actor_rol    = session.get("rol", "—")
    if actor_rol not in ("admin", "empleado") and actor_nombre != "Sistema":
        return
    try:

        conn   = conectar_db()
        cursor = conn.cursor()
        cursor.execute("""
            INSERT INTO auditoria (tipo, actor_id, actor_nombre, actor_rol,
                                   afectado_id, afectado_nombre, detalle)
            VALUES (%s, %s, %s, %s, %s, %s, %s)
        """, (tipo, actor_cui, actor_nombre, actor_rol,
               afectado_id, afectado_nombre, detalle))
        conn.commit()
        conn.close()
    except Exception as e:
        print(f"[LOG ERROR] {e}")


def validar_contrasena(password):
    """
    Valida que la contraseña cumpla con los requisitos mínimos de seguridad:
    - Al menos 8 caracteres
    - Al menos una mayúscula
    - Al menos una minúscula
    - Al menos un número
    - Al menos un carácter especial (ej: @, #, $, !, etc.)
    """
    if len(password) < 8:
        return False, "La contraseña debe tener al menos 8 caracteres"
    if not re.search(r"[A-Z]", password):
        return False, "La contraseña debe tener al menos una mayúscula"
    if not re.search(r"[a-z]", password):
        return False, "La contraseña debe tener al menos una minúscula"
    if not re.search(r"[0-9]", password):
        return False, "La contraseña debe tener al menos un número"
    if not re.search(r"[^A-Za-z0-9]", password):
        return False, "La contraseña debe tener al menos un carácter especial (ej: @, #, $, !)"
    return True, None


def _sumar_meses(anio, mes, cantidad):
    total = (anio * 12) + (mes - 1) + cantidad
    return total // 12, (total % 12) + 1


def periodos_desde_mes_pagado(descripcion):
    if not descripcion:
        return set()

    texto = re.sub(r"\s+", " ", str(descripcion)).strip()
    meses_regex = "|".join(MESES_NOMBRES)
    rango = re.search(
        rf"\b({meses_regex})\s+(\d{{4}})\s*(?:-|\u2013|\u2014|a)\s*({meses_regex})\s+(\d{{4}})\b",
        texto,
        flags=re.IGNORECASE
    )
    if rango:
        mes_inicio = MESES_POR_NOMBRE[rango.group(1).lower()]
        anio_inicio = int(rango.group(2))
        mes_fin = MESES_POR_NOMBRE[rango.group(3).lower()]
        anio_fin = int(rango.group(4))
        periodos = set()
        total_inicio = anio_inicio * 12 + mes_inicio
        total_fin = anio_fin * 12 + mes_fin
        if total_inicio <= total_fin and total_fin - total_inicio < 120:
            for offset in range(total_fin - total_inicio + 1):
                periodos.add(_sumar_meses(anio_inicio, mes_inicio, offset))
        return periodos

    anios = re.findall(r"\b(\d{4})\b", texto)
    if not anios:
        return set()

    anio = int(anios[-1])
    meses = re.findall(rf"\b({meses_regex})\b", texto, flags=re.IGNORECASE)
    return {(anio, MESES_POR_NOMBRE[mes.lower()]) for mes in meses}


def formatear_periodos(periodos):
    ordenados = sorted(periodos)
    return ", ".join(f"{MESES_NOMBRES[mes - 1]} {anio}" for anio, mes in ordenados)


def cargos_pendientes_por_usuario(cursor, cuis):
    resultado = {cui: [] for cui in cuis}
    if not cuis:
        return resultado
    
    placeholders = ",".join(["%s"] * len(cuis))
    cursor.execute(f"SELECT id_cargo, cui_usuario, descripcion, monto, fecha_emision FROM cargos WHERE estado='pendiente' AND cui_usuario IN ({placeholders}) ORDER BY fecha_emision ASC", tuple(cuis))
    for r in cursor.fetchall():
        resultado[r['cui_usuario']].append(r)
    return resultado


def periodos_pagados_por_usuario(cursor, cuis):
    if not cuis:
        return {}

    placeholders = ", ".join(["%s"] * len(cuis))
    cursor.execute(f"""
        SELECT cui_usuario, descripcion
        FROM pagos
        WHERE cui_usuario IN ({placeholders})
    """, tuple(cuis))

    periodos = {str(cui): [] for cui in cuis}
    for pago in cursor.fetchall():
        cui = str(pago["cui_usuario"])
        for anio, mes in sorted(periodos_desde_mes_pagado(pago.get("descripcion"))):
            clave = f"{anio}-{mes}"
            if clave not in periodos.setdefault(cui, []):
                periodos[cui].append(clave)
    return periodos


def enviar_correo_reset(destino, token, nombre):
    gmail_user = GMAIL_USER
    gmail_pwd  = GMAIL_PASSWORD.replace(" ", "") if GMAIL_PASSWORD else None

    if not gmail_user or not gmail_pwd:
        raise ValueError("Variables de entorno GMAIL_USER o GMAIL_PASSWORD no configuradas en Railway.")

    base_url = os.environ.get("BASE_URL", "http://localhost:5000")
    link     = f"{base_url}/reset_password/{token}"
    asunto   = "Recuperación de contraseña — Bodyflex Gym"

    html = f"""
    <!DOCTYPE html>
    <html><head><meta charset="UTF-8"></head>
    <body style="margin:0;padding:0;background:#0f0f0f;font-family:'Helvetica Neue',Arial,sans-serif;">
      <table width="100%" cellpadding="0" cellspacing="0" style="background:#0f0f0f;padding:40px 20px;">
        <tr><td align="center">
          <table width="500" cellpadding="0" cellspacing="0"
                 style="background:#1a1a1a;border-radius:16px;border:1px solid #2e2e2e;overflow:hidden;">
            <tr>
              <td style="background:#141414;padding:28px 36px;border-bottom:1px solid #2e2e2e;">
                <span style="font-weight:900;font-size:22px;color:#f5f5f5;letter-spacing:-0.5px;">
                  BODYFLEX<span style="color:#FF6B00;">GYM</span>
                </span>
              </td>
            </tr>
            <tr>
              <td style="padding:36px;">
                <div style="font-size:36px;margin-bottom:16px;">🔑</div>
                <h1 style="color:#f5f5f5;font-size:22px;font-weight:700;margin:0 0 12px;">
                  Recupera tu contraseña
                </h1>
                <p style="color:#9ca3af;font-size:14px;line-height:1.6;margin:0 0 28px;">
                  Hola <strong style="color:#f5f5f5;">{nombre}</strong>, recibimos una solicitud para
                  restablecer la contraseña de tu cuenta en Bodyflex Gym.
                </p>
                <table cellpadding="0" cellspacing="0" style="margin-bottom:28px;">
                  <tr>
                    <td style="background:#FF6B00;border-radius:10px;">
                      <a href="{link}"
                         style="display:inline-block;padding:14px 32px;color:#fff;text-decoration:none;
                                font-weight:700;font-size:15px;">
                        Restablecer contraseña →
                      </a>
                    </td>
                  </tr>
                </table>
                <div style="background:#222;border:1px solid #2e2e2e;border-radius:10px;padding:16px;margin-bottom:24px;">
                  <p style="color:#9ca3af;font-size:13px;margin:0;">
                    ⏰ Este enlace expira en <strong style="color:#FF6B00;">1 hora</strong>.
                    Si no solicitaste este cambio, ignora este correo.
                  </p>
                </div>
                <p style="color:#555;font-size:11px;margin:0;word-break:break-all;">
                  Si el botón no funciona copia este enlace:<br>
                  <span style="color:#FF6B00;">{link}</span>
                </p>
              </td>
            </tr>
            <tr>
              <td style="background:#141414;padding:20px 36px;border-top:1px solid #2e2e2e;">
                <p style="color:#555;font-size:11px;margin:0;text-align:center;">
                  Bodyflex Gym — Este correo fue generado automáticamente.
                </p>
              </td>
            </tr>
          </table>
        </td></tr>
      </table>
    </body></html>
    """

    texto = f"Hola {nombre},\n\nRestablecer contraseña:\n{link}\n\nExpira en 1 hora.\n\n— Bodyflex Gym"

    msg = MIMEMultipart("alternative")
    msg["Subject"] = asunto
    msg["From"]    = f"Bodyflex Gym <{gmail_user}>"
    msg["To"]      = destino
    msg.attach(MIMEText(texto, "plain"))
    msg.attach(MIMEText(html, "html"))

    contexto_ssl = ssl.create_default_context()
    with smtplib.SMTP_SSL("smtp.gmail.com", 465, context=contexto_ssl) as server:
        server.login(gmail_user, gmail_pwd)
        server.sendmail(gmail_user, destino, msg.as_string())


# ─────────────────────────────────────────────
# REGISTRO
# ─────────────────────────────────────────────

@app.route("/registrar", methods=["POST"])
def registrar():
    nombre     = request.form.get("nombre",     "").strip()
    apellido   = request.form.get("apellido",   "").strip()
    email_raw  = request.form.get("correo",     "").strip()
    password   = request.form.get("password",   "").strip()
    numero_doc = request.form.get("numero_doc", "").strip()
    tipo_doc   = request.form.get("tipo_doc",   "CUI").strip()
    telefono   = request.form.get("telefono",   "").strip()

    # El correo es OPCIONAL — si viene vacío se guarda como NULL
    email = email_raw if email_raw else None

    if not nombre or not apellido or not password or not numero_doc or not telefono:
        flash("Nombre, apellido, documento, teléfono y contraseña son obligatorios", "error")
        return redirect("/registro")

    # Validar email solo si lo proporcionaron
    if email and "@" not in email:
        flash("Correo inválido", "error")
        return redirect("/registro")

    if not numero_doc.isdigit() or len(numero_doc) != 13:
        flash("El CUI/DPI debe tener exactamente 13 dígitos", "error")
        return redirect("/registro")

    if not telefono.isdigit() or len(telefono) != 8:
        flash("El número de teléfono debe tener exactamente 8 dígitos", "error")
        return redirect("/registro")

    if tipo_doc not in ("CUI", "DPI"):
        tipo_doc = "CUI"

    es_valida, msg_error = validar_contrasena(password)
    if not es_valida:
        flash(msg_error, "error")
        return redirect("/registro")

    conn   = conectar_db()
    cursor = conn.cursor()

    # Verificar email duplicado solo si se proporcionó
    if email:
        cursor.execute("SELECT cui FROM usuarios WHERE email=%s", (email,))
        if cursor.fetchone():
            conn.close()
            flash("Ese correo ya está registrado", "error")
            return redirect("/registro")

    # Verificar CUI duplicado
    cursor.execute("SELECT cui FROM usuarios WHERE cui=%s", (int(numero_doc),))
    if cursor.fetchone():
        conn.close()
        flash("Ese CUI/DPI ya está registrado", "error")
        return redirect("/registro")

    password_hash = generate_password_hash(password)
    cursor.execute("""
        INSERT INTO usuarios (cui, tipo_doc, nombre, apellido, email, password, estado, telefono)
        VALUES (%s, %s, %s, %s, %s, %s, 'activo', %s)
    """, (int(numero_doc), tipo_doc, nombre, apellido, email, password_hash, telefono))
    
    cursor.execute("""
        INSERT INTO perfiles (cui_usuario)
        VALUES (%s)
    """, (int(numero_doc),))
    
    conn.commit()
    conn.close()

    flash("Cuenta creada exitosamente. ¡Inicia sesión!", "success")
    return redirect("/login")


# ─────────────────────────────────────────────
# INICIO DE SESIÓN — acepta correo O CUI (13 dígitos)
# ─────────────────────────────────────────────

@app.route("/iniciar", methods=["POST"])
def iniciar():
    identificador = request.form.get("identificador", "").strip()
    password      = request.form.get("password", "")

    if not identificador or not password:
        flash("Completa todos los campos", "error")
        return redirect("/login")

    conn   = conectar_db()
    cursor = conn.cursor(dictionary=True)

    # Detectar si ingresaron CUI (13 dígitos numéricos) o correo
    if identificador.isdigit() and len(identificador) == 13:
        cursor.execute("""
            SELECT u.cui, u.nombre, u.apellido, u.email, u.password, u.estado, r.descripcion AS rol, p.edad
            FROM usuarios u
            JOIN roles r ON u.id_rol = r.id_rol
            LEFT JOIN perfiles p ON u.cui = p.cui_usuario
            WHERE u.cui = %s
        """, (int(identificador),))
    else:
        cursor.execute("""
            SELECT u.cui, u.nombre, u.apellido, u.email, u.password, u.estado, r.descripcion AS rol, p.edad
            FROM usuarios u
            JOIN roles r ON u.id_rol = r.id_rol
            LEFT JOIN perfiles p ON u.cui = p.cui_usuario
            WHERE u.email = %s
        """, (identificador,))

    usuario = cursor.fetchone()

    if not usuario:
        conn.close()
        flash("Identificador o contraseña incorrectos", "error")
        return redirect("/login")

    if usuario["estado"].lower() != "activo":
        conn.close()
        flash("Tu cuenta está inactiva. Contacta al gimnasio.", "error")
        return redirect("/login")

    if not check_password_hash(usuario["password"], password):
        conn.close()
        flash("Identificador o contraseña incorrectos", "error")
        return redirect("/login")

    conn.close()
    session["usuario_id"] = usuario["cui"]
    session["nombre"]     = usuario["nombre"]
    session["rol"]        = usuario["rol"]

    registrar_log("login", "Inició sesión")

    if usuario["rol"] == "admin":
        return redirect("/admin")
    if usuario["rol"] == "empleado":
        return redirect("/empleado")
    if not usuario["edad"]:
        return redirect("/completar_perfil")
    return redirect("/panel")


# ─────────────────────────────────────────────
# PANEL ADMIN
# ─────────────────────────────────────────────

@app.route("/admin")
def admin_panel():
    if "usuario_id" not in session or session.get("rol") != "admin":
        return redirect("/login")

    buscar = request.args.get("buscar")
    filtro = request.args.get("filtro")
    pagina = int(request.args.get("pagina", 1))
    por_pagina = 20

    conn   = conectar_db()
    cursor = conn.cursor(dictionary=True)

    query = """
        SELECT
            u.cui, u.tipo_doc, u.nombre, u.apellido, u.email,
            u.estado, r.descripcion AS rol,
            p.edad, p.peso, p.altura, p.objetivo, u.telefono,
            (SELECT MAX(fecha_vencimiento) FROM pagos WHERE pagos.cui_usuario = u.cui) AS ultimo_vencimiento
        FROM usuarios u
        JOIN roles r ON u.id_rol = r.id_rol
        LEFT JOIN perfiles p ON u.cui = p.cui_usuario
        WHERE u.id_rol = '03'
    """
    params = []

    if buscar:
        query += " AND (u.nombre LIKE %s OR u.apellido LIKE %s)"
        params.extend([f"%{buscar}%", f"%{buscar}%"])

    cursor.execute(query, params)
    usuarios_all = cursor.fetchall()

    cursor.execute("SELECT cui, tipo_doc, nombre, apellido, email, estado FROM usuarios WHERE id_rol='02'")
    empleados = cursor.fetchall()

    periodos_pagados = periodos_pagados_por_usuario(cursor, [u["cui"] for u in usuarios_all])
    cargos_pendientes = cargos_pendientes_por_usuario(cursor, [u["cui"] for u in usuarios_all])

    conn.close()

    fecha_hoy = date.today()

    if filtro == "vencidos":
        usuarios_all = [u for u in usuarios_all if u["ultimo_vencimiento"] and u["ultimo_vencimiento"] < fecha_hoy]
    if filtro == "activos":
        usuarios_all = [u for u in usuarios_all if u["ultimo_vencimiento"] and u["ultimo_vencimiento"] >= fecha_hoy]

    # ── Paginación ──
    total_socios = len(usuarios_all)
    total_paginas = max(1, math.ceil(total_socios / por_pagina))
    pagina = max(1, min(pagina, total_paginas))
    offset = (pagina - 1) * por_pagina
    usuarios = usuarios_all[offset:offset + por_pagina]

    return render_template("admin.html", usuarios=usuarios, empleados=empleados,
                           fecha_hoy=fecha_hoy, precio_mensual=PRECIO_MENSUAL,
                           periodos_pagados=periodos_pagados,
                           cargos_pendientes=cargos_pendientes,
                           pagina=pagina, total_paginas=total_paginas,
                           total_socios=total_socios)


@app.route("/admin/empleados")
def admin_empleados():
    if "usuario_id" not in session or session.get("rol") != "admin":
        return redirect("/login")
    conn = conectar_db()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("SELECT cui, tipo_doc, nombre, apellido, email, estado FROM usuarios WHERE id_rol='02'")
    empleados = cursor.fetchall()
    conn.close()
    return render_template("empleados_admin.html", empleados=empleados, total_empleados=len(empleados))



@app.route("/admin/hacer_admin/<int:cui>", methods=["POST"])
def hacer_admin(cui):
    if "usuario_id" not in session or session.get("rol") != "admin":
        return redirect("/login")
    conn = conectar_db(); cursor = conn.cursor()
    cursor.execute("SELECT nombre, apellido FROM usuarios WHERE cui=%s", (cui,))
    u = cursor.fetchone()
    cursor.execute("UPDATE usuarios SET id_rol='01' WHERE cui=%s", (cui,))
    conn.commit(); conn.close()
    registrar_log("rol", "Promovió a Admin", afectado_id=cui,
                  afectado_nombre=f"{u[0]} {u[1]}" if u else None)
    flash("Usuario promovido a administrador", "success")
    return redirect("/admin")


@app.route("/admin/quitar_admin/<int:cui>", methods=["POST"])
def quitar_admin(cui):
    if "usuario_id" not in session or session.get("rol") != "admin":
        return redirect("/login")
    if cui == session["usuario_id"]:
        flash("No puedes quitarte tu propio rol admin", "error")
        return redirect("/admin")
    conn = conectar_db(); cursor = conn.cursor()
    cursor.execute("SELECT nombre, apellido FROM usuarios WHERE cui=%s", (cui,))
    u = cursor.fetchone()
    cursor.execute("UPDATE usuarios SET id_rol='03' WHERE cui=%s", (cui,))
    conn.commit(); conn.close()
    registrar_log("rol", "Quitó rol Admin → Usuario", afectado_id=cui,
                  afectado_nombre=f"{u[0]} {u[1]}" if u else None)
    flash("Rol admin removido", "success")
    return redirect("/admin")


@app.route("/admin/hacer_empleado/<int:cui>", methods=["POST"])
def hacer_empleado(cui):
    if "usuario_id" not in session or session.get("rol") != "admin":
        return redirect("/login")
    conn = conectar_db(); cursor = conn.cursor()
    cursor.execute("SELECT nombre, apellido FROM usuarios WHERE cui=%s", (cui,))
    u = cursor.fetchone()
    cursor.execute("UPDATE usuarios SET id_rol='02' WHERE cui=%s", (cui,))
    conn.commit(); conn.close()
    registrar_log("rol", "Asignó como Empleado", afectado_id=cui,
                  afectado_nombre=f"{u[0]} {u[1]}" if u else None)
    flash("Usuario asignado como empleado", "success")
    return redirect("/admin")


@app.route("/admin/quitar_empleado/<int:cui>", methods=["POST"])
def quitar_empleado(cui):
    if "usuario_id" not in session or session.get("rol") != "admin":
        return redirect("/login")
    conn = conectar_db(); cursor = conn.cursor()
    cursor.execute("SELECT nombre, apellido FROM usuarios WHERE cui=%s", (cui,))
    u = cursor.fetchone()
    cursor.execute("UPDATE usuarios SET id_rol='03' WHERE cui=%s", (cui,))
    conn.commit(); conn.close()
    registrar_log("rol", "Quitó rol Empleado → Usuario", afectado_id=cui,
                  afectado_nombre=f"{u[0]} {u[1]}" if u else None)
    flash("Rol empleado removido", "success")
    return redirect("/admin")


@app.route("/admin/pagos/<int:cui>")
def ver_pagos(cui):
    if "usuario_id" not in session or session.get("rol") not in ("admin", "empleado"):
        return redirect("/login")

    fecha_inicio = request.args.get("fecha_inicio", "").strip()
    fecha_fin    = request.args.get("fecha_fin", "").strip()

    conn   = conectar_db()
    cursor = conn.cursor(dictionary=True)

    query = """
        SELECT u.nombre, u.apellido,
               p.id_pago, p.monto, p.fecha_pago, p.fecha_vencimiento, p.descripcion
        FROM pagos p
        JOIN usuarios u ON u.cui = p.cui_usuario
        WHERE u.cui = %s
    """
    params = [cui]

    if fecha_inicio:
        query += " AND p.fecha_pago >= %s"
        params.append(fecha_inicio)
    if fecha_fin:
        query += " AND p.fecha_pago <= %s"
        params.append(fecha_fin)

    query += " ORDER BY p.fecha_pago DESC"
    cursor.execute(query, params)
    pagos = cursor.fetchall()

    # Recalculate based on active list
    total_ingresos = sum(float(p["monto"]) for p in pagos)
    total_meses  = int(total_ingresos / PRECIO_MENSUAL)

    nombre_socio = "Sin pagos"
    if pagos:
        nombre_socio = f"{pagos[0]['nombre']} {pagos[0]['apellido']}"
    else:
        # Fetch name if no payment records exist matching filter
        cursor.execute("SELECT nombre, apellido FROM usuarios WHERE cui = %s", (cui,))
        socio = cursor.fetchone()
        if socio:
            nombre_socio = f"{socio['nombre']} {socio['apellido']}"
            
    conn.close()

    return render_template("pagos_admin.html", pagos=pagos, total=total_meses,
                           total_ingresos=total_ingresos,
                           nombre_socio=nombre_socio, cui=cui)


@app.route("/admin/pagos/<int:cui>/excel")
def exportar_pagos_excel(cui):
    if "usuario_id" not in session or session.get("rol") not in ("admin", "empleado"):
        return redirect("/login")

    fecha_inicio = request.args.get("fecha_inicio", "").strip()
    fecha_fin    = request.args.get("fecha_fin", "").strip()

    conn   = conectar_db()
    cursor = conn.cursor(dictionary=True)

    query = """
        SELECT u.nombre, u.apellido,
               p.id_pago, p.monto, p.fecha_pago, p.fecha_vencimiento, p.descripcion
        FROM pagos p
        JOIN usuarios u ON u.cui = p.cui_usuario
        WHERE u.cui = %s
    """
    params = [cui]

    if fecha_inicio:
        query += " AND p.fecha_pago >= %s"
        params.append(fecha_inicio)
    if fecha_fin:
        query += " AND p.fecha_pago <= %s"
        params.append(fecha_fin)

    query += " ORDER BY p.fecha_pago DESC"
    cursor.execute(query, params)
    pagos = cursor.fetchall()
    
    nombre_socio = "Socio"
    if pagos:
        nombre_socio = f"{pagos[0]['nombre']} {pagos[0]['apellido']}"
    else:
        cursor.execute("SELECT nombre, apellido FROM usuarios WHERE cui = %s", (cui,))
        socio = cursor.fetchone()
        if socio:
            nombre_socio = f"{socio['nombre']} {socio['apellido']}"
            
    conn.close()

    headers = ["ID Pago", "Fecha Pago", "Periodo Pagado", "Monto", "Fecha Vencimiento"]
    rows = []
    for p in pagos:
        rows.append([
            p["id_pago"],
            p["fecha_pago"],
            p["descripcion"] or "—",
            float(p["monto"]),
            p["fecha_vencimiento"]
        ])

    excel_bytes = generar_reporte_excel(headers, rows, f"Historial Pagos — {nombre_socio}")
    response = make_response(excel_bytes)
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = f'attachment; filename=pagos_{cui}.xlsx'
    return response


@app.route("/admin/pagos/<int:cui>/pdf")
def exportar_pagos_pdf(cui):
    if "usuario_id" not in session or session.get("rol") not in ("admin", "empleado"):
        return redirect("/login")

    fecha_inicio = request.args.get("fecha_inicio", "").strip()
    fecha_fin    = request.args.get("fecha_fin", "").strip()

    conn   = conectar_db()
    cursor = conn.cursor(dictionary=True)

    query = """
        SELECT u.nombre, u.apellido,
               p.id_pago, p.monto, p.fecha_pago, p.fecha_vencimiento, p.descripcion
        FROM pagos p
        JOIN usuarios u ON u.cui = p.cui_usuario
        WHERE u.cui = %s
    """
    params = [cui]

    if fecha_inicio:
        query += " AND p.fecha_pago >= %s"
        params.append(fecha_inicio)
    if fecha_fin:
        query += " AND p.fecha_pago <= %s"
        params.append(fecha_fin)

    query += " ORDER BY p.fecha_pago DESC"
    cursor.execute(query, params)
    pagos = cursor.fetchall()
    
    nombre_socio = "Socio"
    if pagos:
        nombre_socio = f"{pagos[0]['nombre']} {pagos[0]['apellido']}"
    else:
        cursor.execute("SELECT nombre, apellido FROM usuarios WHERE cui = %s", (cui,))
        socio = cursor.fetchone()
        if socio:
            nombre_socio = f"{socio['nombre']} {socio['apellido']}"
            
    conn.close()

    headers = ["No.", "Fecha Pago", "Mes Pagado", "Vencimiento", "Monto"]
    rows = []
    for idx, p in enumerate(pagos):
        rows.append([
            idx + 1,
            p["fecha_pago"],
            p["descripcion"] or "—",
            p["fecha_vencimiento"],
            float(p["monto"])
        ])

    sub = f"Historial de pagos de membresia de {nombre_socio} (CUI: {cui})"
    if fecha_inicio or fecha_fin:
        sub += f" | Rango: {fecha_inicio or '...'} a {fecha_fin or '...'}"

    pdf_bytes = generar_reporte_pdf(headers, rows, "Historial de Pagos", sub)
    response = make_response(pdf_bytes)
    response.headers['Content-Type'] = 'application/pdf'
    response.headers['Content-Disposition'] = f'attachment; filename=pagos_{cui}.pdf'
    return response



@app.route("/admin/desactivar/<int:cui>", methods=["POST"])
def desactivar_usuario(cui):
    if "usuario_id" not in session or session.get("rol") != "admin":
        return redirect("/login")
    conn = conectar_db(); cursor = conn.cursor()
    cursor.execute("SELECT nombre, apellido FROM usuarios WHERE cui=%s", (cui,))
    u = cursor.fetchone()
    cursor.execute("UPDATE usuarios SET estado='inactivo' WHERE cui=%s", (cui,))
    conn.commit(); conn.close()
    registrar_log("desactivar", "Desactivó la cuenta", afectado_id=cui,
                  afectado_nombre=f"{u[0]} {u[1]}" if u else None)
    flash("Usuario desactivado", "success")
    return redirect("/admin")


@app.route("/admin/reactivar/<int:cui>", methods=["POST"])
def reactivar_usuario(cui):
    if "usuario_id" not in session or session.get("rol") != "admin":
        return redirect("/login")
    conn = conectar_db(); cursor = conn.cursor()
    cursor.execute("SELECT nombre, apellido FROM usuarios WHERE cui=%s", (cui,))
    u = cursor.fetchone()
    cursor.execute("UPDATE usuarios SET estado='activo' WHERE cui=%s", (cui,))
    conn.commit(); conn.close()
    registrar_log("activacion", "Reactivó la cuenta", afectado_id=cui,
                  afectado_nombre=f"{u[0]} {u[1]}" if u else None)
    flash("Usuario reactivado", "success")
    return redirect("/admin")


@app.route("/admin/eliminar_usuario/<int:cui>", methods=["POST"])
def eliminar_usuario(cui):
    if "usuario_id" not in session or session.get("rol") != "admin":
        return redirect("/login")
    
    if cui == session["usuario_id"]:
        flash("No puedes eliminarte a ti mismo", "error")
        return redirect(request.referrer or "/admin")
        
    conn = conectar_db()
    cursor = conn.cursor()
    try:
        cursor.execute("SELECT nombre, apellido, rol FROM usuarios WHERE cui=%s", (cui,))
        u = cursor.fetchone()
        if not u:
            flash("Usuario no encontrado", "error")
            return redirect(request.referrer or "/admin")
            
        nombre_completo = f"{u[0]} {u[1]}"
        rol = u[2]
        
        # Eliminar dependencias
        cursor.execute("DELETE FROM pagos WHERE cui_usuario=%s", (cui,))
        cursor.execute("DELETE FROM recuperar_contra WHERE cui_usuario=%s", (cui,))
        
        try:
            cursor.execute("DELETE FROM asistencia WHERE cui_usuario=%s", (cui,))
        except:
            pass
        try:
            cursor.execute("DELETE FROM metas WHERE cui_usuario=%s", (cui,))
        except:
            pass
            
        cursor.execute("DELETE FROM usuarios WHERE cui=%s", (cui,))
        
        registrar_log("eliminacion", f"Eliminó cuenta de {rol}", afectado_id=cui, afectado_nombre=nombre_completo)
        
        conn.commit()
        flash(f"{'Empleado' if rol == 'empleado' else 'Socio'} eliminado exitosamente", "success")
    except Exception as e:
        conn.rollback()
        flash(f"Error al eliminar: {str(e)}", "error")
    finally:
        conn.close()
        
    return redirect(request.referrer or "/admin")


@app.route("/admin/registrar_pago/<int:cui>", methods=["POST"])
def registrar_pago(cui):
    if "usuario_id" not in session or session.get("rol") not in ("admin", "empleado"):
        return redirect("/login")

    meses_lista = request.form.get("meses_lista", "")
    anio_sel    = request.form.get("anio_sel", str(date.today().year))
    redirect_destino = "/empleado" if session.get("rol") == "empleado" else "/admin"
    conn   = conectar_db()
    cursor = conn.cursor(dictionary=True)

    cursor.execute("SELECT MAX(fecha_vencimiento) AS ultimo FROM pagos WHERE cui_usuario=%s", (cui,))
    resultado = cursor.fetchone()

    cursor.execute("SELECT nombre, apellido FROM usuarios WHERE cui=%s", (cui,))
    socio = cursor.fetchone()

    hoy        = date.today()
    fecha_base = resultado["ultimo"] if resultado["ultimo"] and resultado["ultimo"] >= hoy else hoy

    try:
        anio_i = int(anio_sel)
    except ValueError:
        conn.close()
        flash("Selecciona un anio valido para registrar el pago.", "error")
        return redirect(redirect_destino)

    if meses_lista:
        meses_nums = sorted(set(int(m) for m in meses_lista.split(",") if m.strip().isdigit()))
        if not meses_nums or any(m < 1 or m > 12 for m in meses_nums):
            conn.close()
            flash("Selecciona al menos un mes valido para registrar el pago.", "error")
            return redirect(redirect_destino)

        meses      = len(meses_nums)
        periodos_solicitados = {(anio_i, m) for m in meses_nums}
        nombres_sel = [MESES_NOMBRES[m-1] for m in meses_nums]
        if meses == 1:
            descripcion = f"{nombres_sel[0]} {anio_i}"
        else:
            descripcion = f"{', '.join(nombres_sel[:-1])} y {nombres_sel[-1]} {anio_i}"
        ultimo_mes = max(meses_nums)
        ultimo_anio = anio_i
    else:
        meses  = int(request.form.get("meses", 1))
        mes_i  = fecha_base.month if fecha_base > hoy else hoy.month
        periodos_solicitados = {_sumar_meses(anio_i, mes_i, offset) for offset in range(meses)}
        if meses == 1:
            descripcion = f"{MESES_NOMBRES[mes_i-1]} {anio_i}"
        else:
            mes_fin  = ((mes_i - 1 + meses - 1) % 12) + 1
            anio_fin = anio_i + ((mes_i - 1 + meses - 1) // 12)
            descripcion = f"{MESES_NOMBRES[mes_i-1]} {anio_i} — {MESES_NOMBRES[mes_fin-1]} {anio_fin}"
        ultimo_anio, ultimo_mes = max(periodos_solicitados)

    cursor.execute("SELECT descripcion FROM pagos WHERE cui_usuario=%s", (cui,))
    periodos_registrados = set()
    for pago in cursor.fetchall():
        periodos_registrados.update(periodos_desde_mes_pagado(pago.get("descripcion")))

    duplicados = periodos_solicitados & periodos_registrados
    if duplicados:
        conn.close()
        flash(f"No se puede registrar: {formatear_periodos(duplicados)} ya fue pagado para este socio.", "error")
        return redirect(redirect_destino)

    nueva_fecha = calcular_fecha_vencimiento_dia_3(ultimo_anio, ultimo_mes)
    monto_total = calcular_monto_pago(meses)

    cursor = conn.cursor()
    cursor.execute("INSERT INTO pagos (cui_usuario, fecha_pago, fecha_vencimiento, monto, descripcion) VALUES (%s,%s,%s,%s,%s)",
                   (cui, hoy, nueva_fecha, monto_total, descripcion))
    conn.commit(); conn.close()

    nombre_socio = f"{socio['nombre']} {socio['apellido']}" if socio else "—"
    registrar_log("pago", f"Registró pago de {meses} mes(es) — Q{int(monto_total):,}",
                  afectado_id=cui, afectado_nombre=nombre_socio)

    flash(f"Pago de {meses} mes(es) registrado — Q{int(monto_total):,}", "success")


    return redirect(redirect_destino)


# ─────────────────────────────────────────────
# CARGOS MANUALES
# ─────────────────────────────────────────────

@app.route("/crear_cargo/<int:cui>", methods=["POST"])
def crear_cargo(cui):
    if "usuario_id" not in session or session.get("rol") not in ("admin", "empleado"):
        return redirect("/login")
        
    descripcion = request.form.get("descripcion", "").strip()
    monto = request.form.get("monto", "").strip()
    origen = request.form.get("origen", "/admin").strip()
    
    if not descripcion or not monto:
        flash("Descripción y monto son obligatorios para el cargo", "error")
        return redirect(origen)
        
    conn = conectar_db()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("SELECT nombre, apellido FROM usuarios WHERE cui=%s", (cui,))
    socio = cursor.fetchone()
    
    hoy = date.today()
    cursor.execute("INSERT INTO cargos (cui_usuario, descripcion, monto, fecha_emision, estado) VALUES (%s, %s, %s, %s, 'pendiente')",
                   (cui, descripcion, monto, hoy))
    conn.commit()
    conn.close()
    
    nombre_socio = f"{socio['nombre']} {socio['apellido']}" if socio else "—"
    registrar_log("cargo", f"Creó cargo manual: '{descripcion}' (Q{monto})", afectado_id=cui, afectado_nombre=nombre_socio)
    
    flash(f"Cargo '{descripcion}' creado exitosamente", "success")
    return redirect(origen)


@app.route("/pagar_cargo/<int:id_cargo>", methods=["POST"])
def pagar_cargo(id_cargo):
    if "usuario_id" not in session or session.get("rol") not in ("admin", "empleado"):
        return redirect("/login")
        
    origen = request.form.get("origen", "/admin").strip()
    conn = conectar_db()
    cursor = conn.cursor(dictionary=True)
    
    cursor.execute("SELECT c.cui_usuario, c.monto, c.descripcion, c.estado, u.nombre, u.apellido FROM cargos c JOIN usuarios u ON c.cui_usuario = u.cui WHERE c.id_cargo=%s", (id_cargo,))
    cargo = cursor.fetchone()
    
    if not cargo or cargo['estado'] == 'pagado':
        conn.close()
        flash("Cargo inválido o ya pagado", "error")
        return redirect(origen)
        
    hoy = date.today()
    
    # Obtener el último fecha_vencimiento para no alterarlo
    cursor.execute("SELECT MAX(fecha_vencimiento) as max_venc FROM pagos WHERE cui_usuario=%s", (cargo['cui_usuario'],))
    res_venc = cursor.fetchone()
    vencimiento_actual = res_venc['max_venc'] if res_venc and res_venc['max_venc'] else hoy
    
    # 1. Actualizar estado del cargo
    cursor.execute("UPDATE cargos SET estado='pagado' WHERE id_cargo=%s", (id_cargo,))
    
    # 2. Insertar pago con el id_cargo
    # Ponemos la misma fecha de vencimiento actual para no avanzar su membresía
    cursor.execute("""
        INSERT INTO pagos (cui_usuario, fecha_pago, fecha_vencimiento, monto, descripcion, id_cargo) 
        VALUES (%s, %s, %s, %s, %s, %s)
    """, (cargo['cui_usuario'], hoy, vencimiento_actual, cargo['monto'], f"Cargo: {cargo['descripcion']}", id_cargo))
    
    conn.commit()
    conn.close()
    
    nombre_socio = f"{cargo['nombre']} {cargo['apellido']}"
    registrar_log("pago", f"Pagó cargo manual '{cargo['descripcion']}' (Q{cargo['monto']})", afectado_id=cargo['cui_usuario'], afectado_nombre=nombre_socio)
    
    flash(f"El cargo '{cargo['descripcion']}' ha sido pagado", "success")
    return redirect(origen)# ─────────────────────────────────────────────
# AUDITORÍA
# ─────────────────────────────────────────────

@app.route("/admin/auditoria")
def auditoria():
    if "usuario_id" not in session or session.get("rol") != "admin":
        return redirect("/login")

    buscar       = request.args.get("buscar", "").strip()
    tipo_filtro  = request.args.get("tipo", "").strip()
    fecha_inicio = request.args.get("fecha_inicio", "").strip()
    fecha_fin    = request.args.get("fecha_fin", "").strip()
    pagina       = int(request.args.get("pagina", 1))
    por_pagina   = 30
    offset       = (pagina - 1) * por_pagina

    # Por defecto, si es primera carga sin argumentos, solo mostrar importantes (ocultar login y perfil)
    solo_importantes = request.args.get("solo_importantes", "0")
    if not request.args:
        solo_importantes = "1"

    conn   = conectar_db()
    cursor = conn.cursor(dictionary=True)

    condiciones = []
    params      = []

    if buscar:
        condiciones.append("(actor_nombre LIKE %s OR afectado_nombre LIKE %s OR detalle LIKE %s)")
        params.extend([f"%{buscar}%", f"%{buscar}%", f"%{buscar}%"])

    if tipo_filtro:
        condiciones.append("tipo = %s")
        params.append(tipo_filtro)
    elif solo_importantes == "1":
        condiciones.append("tipo NOT IN ('login', 'perfil')")

    if fecha_inicio:
        condiciones.append("DATE(fecha) >= %s")
        params.append(fecha_inicio)
    if fecha_fin:
        condiciones.append("DATE(fecha) <= %s")
        params.append(fecha_fin)

    where = ("WHERE " + " AND ".join(condiciones)) if condiciones else ""

    cursor.execute(f"SELECT COUNT(*) AS total FROM auditoria {where}", params)
    total_acciones = cursor.fetchone()["total"]
    total_paginas  = max(1, -(-total_acciones // por_pagina))

    cursor.execute(
        f"SELECT * FROM auditoria {where} ORDER BY fecha DESC LIMIT %s OFFSET %s",
        params + [por_pagina, offset]
    )
    auditoria_registros = cursor.fetchall()

    hoy_str = date.today().strftime("%Y-%m-%d")
    cursor.execute("SELECT COUNT(*) AS c FROM auditoria WHERE tipo='pago' AND DATE(fecha)=%s", (hoy_str,))
    pagos_hoy = cursor.fetchone()["c"]

    cursor.execute("SELECT COUNT(*) AS c FROM auditoria WHERE tipo='rol'")
    cambios_rol = cursor.fetchone()["c"]

    cursor.execute("SELECT COUNT(*) AS c FROM auditoria WHERE tipo='desactivar'")
    desactivaciones = cursor.fetchone()["c"]

    conn.close()

    return render_template("auditoria.html",
        auditoria=auditoria_registros,
        total_acciones=total_acciones,
        pagos_hoy=pagos_hoy,
        cambios_rol=cambios_rol,
        desactivaciones=desactivaciones,
        buscar=buscar,
        tipo_filtro=tipo_filtro,
        pagina=pagina,
        total_paginas=total_paginas,
        solo_importantes=solo_importantes,
    )


# Helper function to generate styled Excel report
def generar_reporte_excel(headers, rows, title):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = title[:30]
    ws.views.sheetView[0].showGridLines = True

    font_title = Font(name='Segoe UI', size=15, bold=True, color='FF6B00')
    font_header = Font(name='Segoe UI', size=10, bold=True, color='FFFFFF')
    font_data = Font(name='Segoe UI', size=9.5)
    
    fill_header = PatternFill(start_color='1F1F1F', end_color='1F1F1F', fill_type='solid')
    align_center = Alignment(horizontal='center', vertical='center')
    align_left = Alignment(horizontal='left', vertical='center')
    align_right = Alignment(horizontal='right', vertical='center')

    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )

    # Title Block
    ws.cell(row=1, column=1, value=title).font = font_title
    ws.row_dimensions[1].height = 36
    ws.cell(row=1, column=1).alignment = align_left

    # Headers
    ws.append([]) # row 2 empty
    ws.append(headers) # row 3
    ws.row_dimensions[3].height = 24
    
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=3, column=col_idx)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = thin_border

    # Data Rows
    current_row = 4
    for r in rows:
        row_values = []
        for val in r:
            if isinstance(val, datetime):
                row_values.append(val.strftime('%Y-%m-%d %H:%M:%S'))
            elif isinstance(val, date):
                row_values.append(val.strftime('%Y-%m-%d'))
            else:
                row_values.append(val)
        ws.append(row_values)
        ws.row_dimensions[current_row].height = 20
        
        for col_idx in range(1, len(r) + 1):
            cell = ws.cell(row=current_row, column=col_idx)
            cell.font = font_data
            cell.border = thin_border
            
            val = r[col_idx - 1]
            if isinstance(val, (int, float, decimal.Decimal if 'decimal' in globals() else float)):
                cell.alignment = align_right
                header_name = headers[col_idx - 1].lower()
                if "monto" in header_name or "ingreso" in header_name or "total" in header_name:
                    cell.number_format = '"Q"#,##0'
            elif isinstance(val, (date, datetime)):
                cell.alignment = align_center
            else:
                cell.alignment = align_left
        current_row += 1

    # Auto-width
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val_str = str(cell.value or '')
            if len(val_str) > max_len:
                max_len = len(val_str)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# Helper function to generate multi-page PDF report
def generar_reporte_pdf(headers, rows, title, subtitle):
    from reportlab.lib.pagesizes import letter
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib import colors

    buf = BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=letter,
        leftMargin=36,
        rightMargin=36,
        topMargin=36,
        bottomMargin=36
    )
    story = []

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'DocTitle',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=18,
        textColor=colors.HexColor('#FF6B00'),
        spaceAfter=4
    )
    sub_style = ParagraphStyle(
        'DocSub',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=9,
        textColor=colors.HexColor('#666666'),
        spaceAfter=15
    )
    cell_style = ParagraphStyle(
        'Cell',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=8,
        textColor=colors.HexColor('#333333')
    )
    header_style = ParagraphStyle(
        'HeaderCell',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=8.5,
        textColor=colors.white
    )

    story.append(Paragraph(title, title_style))
    story.append(Paragraph(subtitle, sub_style))

    data = [[Paragraph(h, header_style) for h in headers]]
    for r in rows:
        row_data = []
        for cell_val in r:
            if isinstance(cell_val, datetime):
                txt = cell_val.strftime('%d/%m/%Y %H:%M:%S')
            elif isinstance(cell_val, date):
                txt = cell_val.strftime('%d/%m/%Y')
            elif isinstance(cell_val, (int, float)):
                txt = f"Q{int(cell_val):,}"
            else:
                txt = str(cell_val or '—')
            row_data.append(Paragraph(txt, cell_style))
        data.append(row_data)

    width, height = letter
    usable_width = width - 72
    num_cols = len(headers)
    col_widths = [usable_width / num_cols] * num_cols

    if num_cols == 5: # Auditoria
        col_widths = [usable_width * 0.16, usable_width * 0.14, usable_width * 0.18, usable_width * 0.36, usable_width * 0.16]
    elif num_cols == 5 and headers[0] == "No.": # Pagos socio
        col_widths = [usable_width * 0.08, usable_width * 0.22, usable_width * 0.22, usable_width * 0.30, usable_width * 0.18]

    t = Table(data, colWidths=col_widths, repeatRows=1)
    t.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1F1F1F')),
        ('ALIGN', (0,0), (-1,-1), 'LEFT'),
        ('VALIGN', (0,0), (-1,-1), 'TOP'),
        ('BOTTOMPADDING', (0,0), (-1,-1), 5),
        ('TOPPADDING', (0,0), (-1,-1), 5),
        ('LEFTPADDING', (0,0), (-1,-1), 5),
        ('RIGHTPADDING', (0,0), (-1,-1), 5),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#E5E5E5')),
    ]))
    story.append(t)
    doc.build(story)
    buf.seek(0)
    return buf.read()


@app.route("/admin/auditoria/excel")
def exportar_auditoria_excel():
    if "usuario_id" not in session or session.get("rol") != "admin":
        return redirect("/login")

    buscar       = request.args.get("buscar", "").strip()
    tipo_filtro  = request.args.get("tipo", "").strip()
    fecha_inicio = request.args.get("fecha_inicio", "").strip()
    fecha_fin    = request.args.get("fecha_fin", "").strip()

    conn   = conectar_db()
    cursor = conn.cursor(dictionary=True)

    condiciones = []
    params      = []

    if buscar:
        condiciones.append("(actor_nombre LIKE %s OR afectado_nombre LIKE %s OR detalle LIKE %s)")
        params.extend([f"%{buscar}%", f"%{buscar}%", f"%{buscar}%"])

    if tipo_filtro:
        condiciones.append("tipo = %s")
        params.append(tipo_filtro)

    if fecha_inicio:
        condiciones.append("DATE(fecha) >= %s")
        params.append(fecha_inicio)
    if fecha_fin:
        condiciones.append("DATE(fecha) <= %s")
        params.append(fecha_fin)

    where = ("WHERE " + " AND ".join(condiciones)) if condiciones else ""

    cursor.execute(f"SELECT fecha, tipo, actor_nombre, actor_rol, detalle, afectado_nombre FROM auditoria {where} ORDER BY fecha DESC", params)
    logs = cursor.fetchall()
    conn.close()

    headers = ["Fecha y Hora", "Tipo de Acción", "Realizado por", "Rol Actor", "Detalle", "Socio Afectado"]
    rows = []
    for log in logs:
        rows.append([
            log["fecha"],
            log["tipo"].capitalize(),
            log["actor_nombre"],
            log["actor_rol"],
            log["detalle"],
            log["afectado_nombre"] or "—"
        ])

    excel_bytes = generar_reporte_excel(headers, rows, "Reporte de Auditoria — Bodyflex Gym")
    response = make_response(excel_bytes)
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = 'attachment; filename=reporte_auditoria.xlsx'
    return response


@app.route("/admin/auditoria/pdf")
def exportar_auditoria_pdf():
    if "usuario_id" not in session or session.get("rol") != "admin":
        return redirect("/login")

    buscar       = request.args.get("buscar", "").strip()
    tipo_filtro  = request.args.get("tipo", "").strip()
    fecha_inicio = request.args.get("fecha_inicio", "").strip()
    fecha_fin    = request.args.get("fecha_fin", "").strip()

    conn   = conectar_db()
    cursor = conn.cursor(dictionary=True)

    condiciones = []
    params      = []

    if buscar:
        condiciones.append("(actor_nombre LIKE %s OR afectado_nombre LIKE %s OR detalle LIKE %s)")
        params.extend([f"%{buscar}%", f"%{buscar}%", f"%{buscar}%"])

    if tipo_filtro:
        condiciones.append("tipo = %s")
        params.append(tipo_filtro)

    if fecha_inicio:
        condiciones.append("DATE(fecha) >= %s")
        params.append(fecha_inicio)
    if fecha_fin:
        condiciones.append("DATE(fecha) <= %s")
        params.append(fecha_fin)

    where = ("WHERE " + " AND ".join(condiciones)) if condiciones else ""

    cursor.execute(f"SELECT fecha, tipo, actor_nombre, actor_rol, detalle, afectado_nombre FROM auditoria {where} ORDER BY fecha DESC", params)
    logs = cursor.fetchall()
    conn.close()

    headers = ["Fecha", "Acción", "Realizado por", "Detalle", "Afectado"]
    rows = []
    for log in logs:
        rows.append([
            log["fecha"],
            log["tipo"].capitalize(),
            f"{log['actor_nombre']} ({log['actor_rol']})",
            log["detalle"],
            log["afectado_nombre"] or "—"
        ])

    filtro_txt = f"Filtros: Busqueda: '{buscar or 'Todas'}' | Tipo: '{tipo_filtro or 'Todos'}'"
    if fecha_inicio or fecha_fin:
        filtro_txt += f" | Rango: {fecha_inicio or '...'} a {fecha_fin or '...'}"

    pdf_bytes = generar_reporte_pdf(headers, rows, "Reporte de Auditoría", filtro_txt)
    response = make_response(pdf_bytes)
    response.headers['Content-Type'] = 'application/pdf'
    response.headers['Content-Disposition'] = 'attachment; filename=reporte_auditoria.pdf'
    return response


@app.route("/admin/auditoria/limpiar", methods=["POST"])
def limpiar_auditoria():
    if "usuario_id" not in session or session.get("rol") != "admin":
        return redirect("/login")

    conn = conectar_db()
    cursor = conn.cursor()
    cursor.execute("DELETE FROM auditoria WHERE actor_rol IN ('user', 'socio')")
    conn.commit()
    conn.close()

    flash("Se eliminaron todos los registros de auditoría de usuarios regulares exitosamente.", "success")
    return redirect("/admin/auditoria")


@app.route("/admin/reportes")
def admin_reportes():
    if "usuario_id" not in session or session.get("rol") != "admin":
        return redirect("/login")

    conn   = conectar_db()
    cursor = conn.cursor(dictionary=True)

    # ── Métricas de ingresos mensuales (últimos 12 meses) ──
    cursor.execute("""
        SELECT YEAR(fecha_pago) AS anio, MONTH(fecha_pago) AS mes,
               SUM(monto) AS total
        FROM pagos
        WHERE fecha_pago >= DATE_SUB(CURDATE(), INTERVAL 12 MONTH)
        GROUP BY YEAR(fecha_pago), MONTH(fecha_pago)
        ORDER BY anio, mes
    """)
    ingresos_raw = cursor.fetchall()

    cursor.execute("SELECT COALESCE(SUM(monto), 0) AS total FROM pagos")
    total_ingresos_global = cursor.fetchone()["total"]

    cursor.execute("""
        SELECT COALESCE(SUM(monto), 0) AS total FROM pagos
        WHERE YEAR(fecha_pago) = YEAR(CURDATE()) AND MONTH(fecha_pago) = MONTH(CURDATE())
    """)
    ingresos_mes_actual = cursor.fetchone()["total"]

    conn.close()

    ingresos_labels = [f"{MESES_NOMBRES[r['mes']-1][:3]} {r['anio']}" for r in ingresos_raw]
    ingresos_data   = [float(r['total']) for r in ingresos_raw]

    fecha_hoy = date.today()

    return render_template("reportes.html", 
                           total_ingresos_global=total_ingresos_global,
                           ingresos_mes_actual=ingresos_mes_actual,
                           ingresos_labels=ingresos_labels,
                           ingresos_data=ingresos_data,
                           fecha_hoy=fecha_hoy)


@app.route("/admin/auditoria/borrar_log/<int:id_log>", methods=["POST"])
def borrar_log_individual(id_log):
    if "usuario_id" not in session or session.get("rol") != "admin":
        return redirect("/login")

    conn = conectar_db()
    cursor = conn.cursor()
    cursor.execute("DELETE FROM auditoria WHERE id_log = %s", (id_log,))
    conn.commit()
    conn.close()

    flash("Registro de auditoría eliminado exitosamente.", "success")
    return redirect(request.referrer or "/admin/auditoria")


@app.route("/admin/auditoria/borrar_usuario/<int:cui>", methods=["POST"])
def borrar_logs_usuario(cui):
    if "usuario_id" not in session or session.get("rol") != "admin":
        return redirect("/login")

    conn = conectar_db()
    cursor = conn.cursor()
    cursor.execute("DELETE FROM auditoria WHERE actor_id = %s OR afectado_id = %s", (cui, cui))
    conn.commit()
    conn.close()

    flash("Se eliminaron todos los registros de auditoría asociados a este usuario.", "success")
    return redirect(request.referrer or "/admin/auditoria")



# ─────────────────────────────────────────────
# ADMIN — Restablecer contraseña de un socio
# ─────────────────────────────────────────────

@app.route("/admin/reset_pass/<int:cui>", methods=["POST"])
def admin_reset_pass(cui):
    """El admin genera una nueva contraseña temporal para un socio sin correo."""
    if "usuario_id" not in session or session.get("rol") != "admin":
        return redirect("/login")

    nueva_pass = request.form.get("nueva_pass", "").strip()

    es_valida, msg_error = validar_contrasena(nueva_pass)
    if not es_valida:
        flash(msg_error, "error")
        return redirect("/admin")

    conn = conectar_db(); cursor = conn.cursor(dictionary=True)
    cursor.execute("SELECT nombre, apellido FROM usuarios WHERE cui=%s", (cui,))
    u = cursor.fetchone()
    if not u:
        conn.close()
        flash("Socio no encontrado", "error")
        return redirect("/admin")

    nuevo_hash = generate_password_hash(nueva_pass)
    cursor = conn.cursor()
    cursor.execute("UPDATE usuarios SET password=%s WHERE cui=%s", (nuevo_hash, cui))
    conn.commit(); conn.close()

    registrar_log("perfil", f"Admin restableció contraseña temporalmente",
                  afectado_id=cui, afectado_nombre=f"{u['nombre']} {u['apellido']}")
    flash(f"Contraseña restablecida para {u['nombre']} {u['apellido']}", "success")
    return redirect("/admin")


# ─────────────────────────────────────────────
# CAMBIAR CONTRASEÑA
# ─────────────────────────────────────────────

@app.route("/cambiar_password", methods=["GET", "POST"])
def cambiar_password():
    if "usuario_id" not in session:
        return redirect("/login")

    if request.method == "GET":
        return render_template("cambiar_password.html")

    actual    = request.form.get("password_actual", "").strip()
    nueva     = request.form.get("password_nueva", "").strip()
    confirmar = request.form.get("password_confirmar", "").strip()

    if not actual or not nueva or not confirmar:
        flash("Todos los campos son obligatorios", "error")
        return redirect("/cambiar_password")

    if nueva != confirmar:
        flash("Las contraseñas nuevas no coinciden", "error")
        return redirect("/cambiar_password")

    es_valida, msg_error = validar_contrasena(nueva)
    if not es_valida:
        flash(msg_error, "error")
        return redirect("/cambiar_password")

    if nueva == actual:
        flash("La nueva contraseña debe ser diferente a la actual", "error")
        return redirect("/cambiar_password")

    conn   = conectar_db()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("SELECT password FROM usuarios WHERE cui=%s", (session["usuario_id"],))
    usuario = cursor.fetchone()

    if not usuario or not check_password_hash(usuario["password"], actual):
        conn.close()
        flash("La contraseña actual es incorrecta", "error")
        return redirect("/cambiar_password")

    nuevo_hash = generate_password_hash(nueva)
    cursor = conn.cursor()
    cursor.execute("UPDATE usuarios SET password=%s WHERE cui=%s",
                   (nuevo_hash, session["usuario_id"]))
    conn.commit()
    conn.close()

    registrar_log("perfil", "Cambió su contraseña")
    flash("Contraseña actualizada correctamente ✓", "success")

    rol = session.get("rol")
    if rol == "admin":     return redirect("/admin")
    if rol == "empleado":  return redirect("/empleado")
    return redirect("/panel")


# ─────────────────────────────────────────────
# RECUPERAR CONTRASEÑA (olvidada)
# ─────────────────────────────────────────────

@app.route("/recuperar_contra", methods=["GET", "POST"])
def recuperar_contra_form():
    if request.method == "GET":
        return render_template("recuperar_contra.html")

    correo = request.form.get("correo", "").strip().lower()

    if not correo or "@" not in correo:
        flash("Ingresa un correo válido", "error")
        return redirect("/recuperar_contra")

    conn   = conectar_db()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("SELECT cui, nombre, email FROM usuarios WHERE email=%s AND estado='activo'", (correo,))
    usuario = cursor.fetchone()

    if usuario:
        token  = secrets.token_urlsafe(48)
        expira = datetime.now() + timedelta(hours=1)

        cursor = conn.cursor()
        cursor.execute("UPDATE recuperar_contra SET usado=1 WHERE cui_usuario=%s AND usado=0", (usuario["cui"],))
        cursor.execute("""
            INSERT INTO recuperar_contra (cui_usuario, token, expira)
            VALUES (%s, %s, %s)
        """, (usuario["cui"], token, expira))
        conn.commit()
        conn.close()

        try:
            enviar_correo_reset(usuario["email"], token, usuario["nombre"])
        except Exception as e:
            print(f"[EMAIL ERROR] {e}")
            flash(f"Error al enviar el correo: {str(e)}", "error")
            return redirect("/recuperar_contra")
    else:
        conn.close()

    flash("Si ese correo está registrado, recibirás un enlace en los próximos minutos.", "success")
    return redirect("/recuperar_contra")


@app.route("/reset_password/<token>", methods=["GET", "POST"])
def reset_password_form(token):
    if request.method == "GET":
        conn   = conectar_db()
        cursor = conn.cursor(dictionary=True)
        cursor.execute("""
            SELECT * FROM recuperar_contra
            WHERE token=%s AND usado=0 AND expira > NOW()
        """, (token,))
        reset = cursor.fetchone()
        conn.close()

        if not reset:
            flash("El enlace es inválido o ya expiró. Solicita uno nuevo.", "error")
            return redirect("/recuperar_contra")

        return render_template("reset_password.html", token=token)

    nueva     = request.form.get("password_nueva", "").strip()
    confirmar = request.form.get("password_confirmar", "").strip()

    if not nueva or not confirmar:
        flash("Completa todos los campos", "error")
        return redirect(f"/reset_password/{token}")

    if nueva != confirmar:
        flash("Las contraseñas no coinciden", "error")
        return redirect(f"/reset_password/{token}")

    es_valida, msg_error = validar_contrasena(nueva)
    if not es_valida:
        flash(msg_error, "error")
        return redirect(f"/reset_password/{token}")

    conn   = conectar_db()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("""
        SELECT * FROM recuperar_contra
        WHERE token=%s AND usado=0 AND expira > NOW()
    """, (token,))
    reset = cursor.fetchone()

    if not reset:
        conn.close()
        flash("El enlace expiró. Solicita uno nuevo.", "error")
        return redirect("/recuperar_contra")

    nuevo_hash = generate_password_hash(nueva)
    cursor = conn.cursor()
    cursor.execute("UPDATE usuarios SET password=%s WHERE cui=%s",
                   (nuevo_hash, reset["cui_usuario"]))
    cursor.execute("UPDATE recuperar_contra SET usado=1 WHERE token=%s", (token,))
    conn.commit()
    conn.close()

    flash("¡Contraseña restablecida! Ya puedes iniciar sesión.", "success")
    return redirect("/login")


# ─────────────────────────────────────────────
# PANEL EMPLEADO
# ─────────────────────────────────────────────

@app.route("/empleado")
def empleado_panel():
    if "usuario_id" not in session or session.get("rol") not in ("admin", "empleado"):
        return redirect("/login")

    buscar = request.args.get("buscar", "")
    conn   = conectar_db()
    cursor = conn.cursor(dictionary=True)

    query = """
        SELECT u.cui, u.tipo_doc, u.nombre, u.apellido, u.estado,
               p.edad, p.peso, p.altura, p.objetivo, u.telefono,
               (SELECT MAX(fecha_vencimiento) FROM pagos WHERE cui_usuario = u.cui) AS ultimo_vencimiento
        FROM usuarios u
        JOIN roles r ON u.id_rol = r.id_rol
        LEFT JOIN perfiles p ON u.cui = p.cui_usuario
        WHERE u.id_rol = '03'
    """
    params = []
    if buscar:
        query += " AND (u.nombre LIKE %s OR u.apellido LIKE %s OR CAST(u.cui AS CHAR) LIKE %s)"
        params.extend([f"%{buscar}%", f"%{buscar}%", f"%{buscar}%"])

    query += " ORDER BY u.nombre"
    cursor.execute(query, params)
    usuarios = cursor.fetchall()
    cui_list = [u["cui"] for u in usuarios]
    periodos_pagados = periodos_pagados_por_usuario(cursor, cui_list)
    cargos_pendientes = cargos_pendientes_por_usuario(cursor, cui_list)
    conn.close()

    return render_template("empleado.html", usuarios=usuarios, fecha_hoy=date.today(),
                           precio_mensual=PRECIO_MENSUAL, buscar=buscar,
                           periodos_pagados=periodos_pagados, cargos_pendientes=cargos_pendientes)


# ─────────────────────────────────────────────
# PANEL USUARIO
# ─────────────────────────────────────────────

@app.route("/panel")
def panel():
    if "usuario_id" not in session:
        return redirect("/login")
    conn   = conectar_db()
    cursor = conn.cursor(dictionary=True)

    cursor.execute("""
        SELECT u.nombre, u.apellido, u.email, u.fecha_registro, u.telefono,
               p.edad, p.peso, p.altura, p.objetivo
        FROM usuarios u
        LEFT JOIN perfiles p ON u.cui = p.cui_usuario
        WHERE u.cui=%s
    """, (session["usuario_id"],))
    perfil = cursor.fetchone()

    imc = None
    imc_categoria = None
    imc_color = None
    imc_consejo = None
    if perfil and perfil["peso"] and perfil["altura"] and float(perfil["altura"]) > 0:
        peso_kg = float(perfil["peso"]) * 0.453592
        altura_m = float(perfil["altura"])
        imc = round(peso_kg / (altura_m ** 2), 1)
        if imc < 18.5:
            imc_categoria = "Bajo peso"
            imc_color = "blue"
            imc_consejo = "Tu peso está por debajo del rango saludable. Considera aumentar tu ingesta calórica con alimentos nutritivos."
        elif imc < 25:
            imc_categoria = "Peso normal"
            imc_color = "green"
            imc_consejo = "¡Excelente! Tu peso está en el rango saludable. Mantén tus hábitos de ejercicio y alimentación."
        elif imc < 30:
            imc_categoria = "Sobrepeso"
            imc_color = "yellow"
            imc_consejo = "Estás ligeramente por encima del rango saludable. El ejercicio regular y una dieta balanceada te ayudarán."
        else:
            imc_categoria = "Obesidad"
            imc_color = "red"
            imc_consejo = "Te recomendamos consultar con un especialista para un plan personalizado de ejercicio y nutrición."

    cursor.execute("""
        SELECT COUNT(*) AS total_pagos,
               COALESCE(SUM(monto), 0) AS total_pagado,
               MAX(fecha_vencimiento) AS vencimiento
        FROM pagos WHERE cui_usuario=%s
    """, (session["usuario_id"],))
    stats = cursor.fetchone()

    meses_miembro = 0
    if perfil and perfil["fecha_registro"]:
        hoy = date.today()
        reg = perfil["fecha_registro"]
        if hasattr(reg, 'date'):
            reg = reg.date()
        meses_miembro = (hoy.year - reg.year) * 12 + (hoy.month - reg.month)

    cursor.execute("""
        SELECT id_pago, fecha_pago, fecha_vencimiento, monto, descripcion
        FROM pagos WHERE cui_usuario=%s
        ORDER BY fecha_pago DESC
    """, (session["usuario_id"],))
    historial_pagos = cursor.fetchall()

    hoy_date = date.today()
    cursor.execute("""
        SELECT YEAR(fecha_pago) as anio, MONTH(fecha_pago) as mes
        FROM pagos WHERE cui_usuario=%s
        ORDER BY fecha_pago DESC
    """, (session["usuario_id"],))
    pagos_meses = cursor.fetchall()
    streak = 0
    if pagos_meses:
        seen = set((r["anio"], r["mes"]) for r in pagos_meses)
        check_year, check_month = hoy_date.year, hoy_date.month
        for _ in range(120):
            if (check_year, check_month) in seen:
                streak += 1
                if check_month == 1:
                    check_month = 12; check_year -= 1
                else:
                    check_month -= 1
            else:
                break

    conn.close()
    return render_template("panel.html", perfil=perfil,
                           imc=imc, imc_categoria=imc_categoria,
                           imc_color=imc_color, imc_consejo=imc_consejo,
                           stats=stats, meses_miembro=meses_miembro,
                           historial_pagos=historial_pagos,
                           streak=streak)


@app.route("/completar_perfil")
def completar_perfil():
    if "usuario_id" not in session:
        return redirect("/login")
    return render_template("perfil.html")


@app.route("/guardar_perfil", methods=["POST"])
def guardar_perfil():
    if "usuario_id" not in session:
        return redirect("/login")
    edad = request.form.get("edad", "").strip()
    peso = request.form.get("peso", "").strip() or None
    altura = request.form.get("altura", "").strip() or None
    objetivo = request.form.get("objetivo", "").strip()
    conn = conectar_db(); cursor = conn.cursor()
    cursor.execute("UPDATE perfiles SET edad=%s, peso=%s, altura=%s, objetivo=%s WHERE cui_usuario=%s",
                   (edad, peso, altura, objetivo, session["usuario_id"]))
    conn.commit(); conn.close()
    registrar_log("perfil", f"Completó perfil — Objetivo: {objetivo}")
    return redirect("/panel")


@app.route("/actualizar_info", methods=["POST"])
def actualizar_info():
    if "usuario_id" not in session:
        return redirect("/login")
    nombre   = request.form.get("nombre")
    apellido = request.form.get("apellido")
    email    = request.form.get("email") or None   # Vacío → NULL
    peso     = request.form.get("peso")
    telefono = request.form.get("telefono", "").strip()

    if not telefono or not telefono.isdigit() or len(telefono) != 8:
        flash("El número de teléfono debe tener exactamente 8 dígitos", "error")
        return redirect("/panel")

    conn = conectar_db(); cursor = conn.cursor()
    if nombre and apellido:
        cursor.execute("UPDATE usuarios SET nombre=%s, apellido=%s, email=%s, telefono=%s WHERE cui=%s",
                       (nombre, apellido, email, telefono, session["usuario_id"]))
    if peso:
        cursor.execute("UPDATE perfiles SET peso=%s WHERE cui_usuario=%s", (peso, session["usuario_id"]))
    conn.commit(); conn.close()
    if nombre: session["nombre"] = nombre
    registrar_log("perfil", "Actualizó su información personal")
    flash("Información actualizada", "success")
    return redirect("/panel")


@app.route("/actualizar_objetivo", methods=["POST"])
def actualizar_objetivo():
    if "usuario_id" not in session:
        return redirect("/login")
    objetivo = request.form.get("objetivo")
    if objetivo:
        conn = conectar_db(); cursor = conn.cursor()
        cursor.execute("UPDATE perfiles SET objetivo=%s WHERE cui_usuario=%s",
                       (objetivo, session["usuario_id"]))
        conn.commit(); conn.close()
        registrar_log("perfil", f"Cambió su objetivo a: {objetivo}")
        flash("Objetivo actualizado correctamente", "success")
    return redirect("/panel")


# ─────────────────────────────────────────────
# RECIBO PDF
# ─────────────────────────────────────────────

@app.route("/recibo/<int:id_pago>")
def generar_recibo(id_pago):
    if "usuario_id" not in session:
        return redirect("/login")

    conn   = conectar_db()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("""
        SELECT p.id_pago, p.monto, p.fecha_pago, p.fecha_vencimiento,
               u.cui, u.tipo_doc, u.nombre, u.apellido, u.email
        FROM pagos p
        JOIN usuarios u ON u.cui = p.cui_usuario
        WHERE p.id_pago = %s
    """, (id_pago,))
    pago = cursor.fetchone()
    conn.close()

    if not pago:
        flash("Pago no encontrado", "error")
        return redirect("/panel")

    es_admin_emp = session.get("rol") in ("admin", "empleado")
    es_dueno     = pago["cui"] == session.get("usuario_id")
    if not es_admin_emp and not es_dueno:
        return redirect("/panel")

    buf = BytesIO()
    c = canvas.Canvas(buf, pagesize=letter)
    width, height = letter

    negro     = HexColor("#0d0d0d")
    naranja   = HexColor("#FF6B00")
    gris_dark = HexColor("#333333")
    gris      = HexColor("#666666")
    blanco    = HexColor("#ffffff")
    verde     = HexColor("#22c55e")

    c.setFillColor(negro)
    c.rect(0, height - 140, width, 140, fill=True, stroke=False)

    logo_path = os.path.join(app.static_folder, 'logo.png')
    if os.path.exists(logo_path):
        c.drawImage(logo_path, 50, height - 95, width=55, height=55,
                    preserveAspectRatio=True, mask='auto')
    c.setFillColor(blanco)
    c.setFont("Helvetica-Bold", 28)
    c.drawString(115, height - 60, "BODYFLEX")
    c.setFillColor(naranja)
    c.drawString(115 + c.stringWidth("BODYFLEX", "Helvetica-Bold", 28), height - 60, "GYM")

    c.setFillColor(HexColor("#aaaaaa"))
    c.setFont("Helvetica", 10)
    c.drawString(115, height - 80, "Recibo de pago de membresía")

    c.setFillColor(naranja)
    c.setFont("Helvetica-Bold", 12)
    c.drawRightString(width - 50, height - 55, f"RECIBO #{pago['id_pago']:04d}")

    c.setFillColor(HexColor("#aaaaaa"))
    c.setFont("Helvetica", 10)
    fecha_str = pago['fecha_pago'].strftime('%d/%m/%Y') if pago['fecha_pago'] else 'N/A'
    c.drawRightString(width - 50, height - 75, f"Fecha: {fecha_str}")

    c.setStrokeColor(naranja)
    c.setLineWidth(3)
    c.line(50, height - 150, width - 50, height - 150)

    y = height - 195
    c.setFillColor(gris)
    c.setFont("Helvetica", 9)
    c.drawString(50, y + 15, "DATOS DEL SOCIO")

    c.setFillColor(negro)
    c.setFont("Helvetica-Bold", 13)
    c.drawString(50, y - 5, f"{pago['nombre']} {pago['apellido']}")

    c.setFillColor(gris_dark)
    c.setFont("Helvetica", 10)
    if pago['email']:
        c.drawString(50, y - 22, f"Correo: {pago['email']}")
        c.drawString(50, y - 38, f"No. Socio: {pago['tipo_doc']}: {pago['cui']}")
    else:
        c.drawString(50, y - 22, f"No. Socio: {pago['tipo_doc']}: {pago['cui']}")

    y = height - 290
    c.setFillColor(HexColor("#f5f5f5"))
    c.rect(50, y - 5, width - 100, 28, fill=True, stroke=False)

    c.setFillColor(gris)
    c.setFont("Helvetica-Bold", 9)
    c.drawString(60, y + 3, "PAGO")
    c.drawString(280, y + 3, "FECHA PAGO")
    c.drawString(390, y + 3, "VENCIMIENTO")
    c.drawRightString(width - 60, y + 3, "MONTO")

    y -= 32
    c.setFillColor(negro)
    c.setFont("Helvetica", 11)
    meses = int(float(pago['monto']) / PRECIO_MENSUAL)
    concepto = f"Membresía ({meses} {'mes' if meses == 1 else 'meses'})"
    c.drawString(60, y + 3, concepto)

    c.setFont("Helvetica", 10)
    c.drawString(280, y + 3, fecha_str)
    venc_str = pago['fecha_vencimiento'].strftime('%d/%m/%Y') if pago['fecha_vencimiento'] else 'N/A'
    c.drawString(390, y + 3, venc_str)

    c.setFont("Helvetica-Bold", 12)
    c.setFillColor(verde)
    c.drawRightString(width - 60, y + 3, f"Q{pago['monto']:.2f}")

    y -= 15
    c.setStrokeColor(HexColor("#e0e0e0"))
    c.setLineWidth(0.5)
    c.line(50, y, width - 50, y)

    y -= 30
    c.setFillColor(negro)
    c.setFont("Helvetica-Bold", 10)
    c.drawString(60, y + 3, "TOTAL")
    c.setFont("Helvetica-Bold", 16)
    c.drawRightString(width - 60, y + 3, f"Q{pago['monto']:.2f}")

    y -= 70
    c.setStrokeColor(verde)
    c.setFillColor(verde)
    c.setLineWidth(2)
    c.roundRect(width / 2 - 60, y - 5, 120, 35, 6, fill=False, stroke=True)
    c.setFont("Helvetica-Bold", 18)
    c.drawCentredString(width / 2, y + 5, "PAGADO")

    c.setFillColor(HexColor("#f8f8f8"))
    c.rect(0, 0, width, 60, fill=True, stroke=False)
    c.setStrokeColor(HexColor("#e0e0e0"))
    c.setLineWidth(0.5)
    c.line(0, 60, width, 60)
    c.setFillColor(gris)
    c.setFont("Helvetica", 8)
    c.drawCentredString(width / 2, 35, "Bodyflex Gym — Sistema de Gestión de Membresías")
    c.drawCentredString(width / 2, 22, f"Recibo generado automáticamente · {date.today().strftime('%d/%m/%Y')}")

    c.showPage()
    c.save()
    buf.seek(0)

    response = make_response(buf.read())
    response.headers['Content-Type'] = 'application/pdf'
    response.headers['Content-Disposition'] = f'inline; filename=recibo_{pago["id_pago"]:04d}.pdf'
    return response


# ─────────────────────────────────────────────
# RUTA DE PRUEBA DE CORREO
# ─────────────────────────────────────────────

@app.route("/test_email")
def test_email():
    if "usuario_id" not in session or session.get("rol") != "admin":
        return redirect("/login")
    try:
        gmail_user = GMAIL_USER
        gmail_pwd  = GMAIL_PASSWORD.replace(" ", "") if GMAIL_PASSWORD else None

        if not gmail_user or not gmail_pwd:
            raise ValueError("GMAIL_USER o GMAIL_PASSWORD no están configuradas en Railway.")

        contexto_ssl = ssl.create_default_context()
        with smtplib.SMTP_SSL("smtp.gmail.com", 465, context=contexto_ssl) as server:
            server.login(gmail_user, gmail_pwd)
            msg = MIMEText("✅ Correo de prueba desde Bodyflex Gym — configuración correcta.")
            msg["Subject"] = "Prueba de correo — Bodyflex Gym"
            msg["From"]    = gmail_user
            msg["To"]      = gmail_user
            server.sendmail(gmail_user, gmail_user, msg.as_string())

        return f"""
        <div style='font-family:sans-serif;padding:40px;background:#0f0f0f;color:#f5f5f5;min-height:100vh;'>
            <h2 style='color:#22c55e;'>✅ Correo enviado correctamente</h2>
            <p>Revisa tu bandeja de entrada en <strong>{gmail_user}</strong></p>
            <a href='/admin' style='color:#FF6B00;'>← Volver al panel</a>
        </div>"""
    except Exception as e:
        pwd_len = len(GMAIL_PASSWORD.replace(" ","")) if GMAIL_PASSWORD else 0
        return f"""
        <div style='font-family:sans-serif;padding:40px;background:#0f0f0f;color:#f5f5f5;min-height:100vh;'>
            <h2 style='color:#ef4444;'>❌ Error al enviar</h2>
            <p style='background:#1a1a1a;padding:16px;border-radius:8px;color:#ef4444;font-family:monospace;'>{str(e)}</p>
            <a href='/admin' style='color:#FF6B00;'>← Volver al panel</a>
        </div>"""


# ─────────────────────────────────────────────
# RUTAS SIMPLES
# ─────────────────────────────────────────────

@app.route("/login")
def login():
    return render_template("login.html")

@app.route("/registro")
def registro():
    return render_template("registro.html")

@app.route("/logout")
def logout():
    registrar_log("login", "Cerró sesión")
    session.clear()
    return redirect("/login")

@app.route("/")
def inicio():
    return render_template("inicio.html")


@app.route("/contrato/<int:cui>")
def generar_contrato_pdf(cui):
    if "usuario_id" not in session:
        return redirect("/login")
        
    es_admin_emp = session.get("rol") in ("admin", "empleado")
    es_dueno     = cui == session.get("usuario_id")
    if not es_admin_emp and not es_dueno:
        return redirect("/panel")
        
    conn = conectar_db()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("SELECT nombre, apellido, email, tipo_doc, cui, fecha_registro, telefono FROM usuarios WHERE cui = %s", (cui,))
    usuario = cursor.fetchone()
    conn.close()
    
    if not usuario:
        flash("Socio no encontrado", "error")
        return redirect("/admin" if es_admin_emp else "/panel")
        
    buf = BytesIO()
    from reportlab.lib.pagesizes import letter
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib import colors
    
    doc = SimpleDocTemplate(
        buf,
        pagesize=letter,
        leftMargin=54,
        rightMargin=54,
        topMargin=54,
        bottomMargin=54
    )
    story = []
    
    styles = getSampleStyleSheet()
    
    title_style = ParagraphStyle(
        'DocTitle',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=14,
        alignment=1,
        spaceAfter=15,
        textColor=colors.HexColor('#1A1A1A')
    )
    section_style = ParagraphStyle(
        'DocSection',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=11,
        spaceBefore=10,
        spaceAfter=5,
        textColor=colors.HexColor('#FF6B00')
    )
    body_style = ParagraphStyle(
        'DocBody',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=9.5,
        leading=14,
        spaceAfter=8,
        textColor=colors.HexColor('#333333')
    )
    bullet_style = ParagraphStyle(
        'DocBullet',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=9.5,
        leading=14,
        leftIndent=15,
        spaceAfter=4,
        textColor=colors.HexColor('#333333')
    )
    
    story.append(Paragraph("CONTRATO DE MEMBRESÍA Y REGLAMENTO DE CONVIVENCIA", title_style))
    story.append(Paragraph("<b>BODYFLEX GYM</b>", ParagraphStyle('Sub', parent=title_style, fontSize=12, spaceAfter=20)))
    
    reg_fecha = usuario["fecha_registro"].strftime("%d/%m/%Y") if usuario["fecha_registro"] else "—"
    detalles_texto = f"""
    <b>DATOS DEL SOCIO:</b><br/>
    <b>Nombre Completo:</b> {usuario['nombre']} {usuario['apellido']}<br/>
    <b>Identificación ({usuario['tipo_doc']}):</b> {usuario['cui']}<br/>
    <b>Teléfono:</b> {usuario['telefono'] or '—'}<br/>
    <b>Correo Electrónico:</b> {usuario['email'] or '—'}<br/>
    <b>Fecha de Registro:</b> {reg_fecha}
    """
    
    t_data = [[Paragraph(detalles_texto, body_style)]]
    t = Table(t_data, colWidths=[letter[0] - 108])
    t.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,-1), colors.HexColor('#F9F9F9')),
        ('BORDER', (0,0), (-1,-1), 1, colors.HexColor('#E5E5E5')),
        ('PADDING', (0,0), (-1,-1), 12),
    ]))
    story.append(t)
    story.append(Spacer(1, 15))
    
    story.append(Paragraph("DECLARACIONES Y CONDICIONES DEL SERVICIO", section_style))
    story.append(Paragraph(
        "Por medio del presente documento, el Socio arriba mencionado acepta y se adhiere formalmente al reglamento de "
        "convivencia y condiciones de membresía de <b>Bodyflex Gym</b>, de conformidad con las siguientes cláusulas:",
        body_style
    ))
    
    story.append(Paragraph("<b>CLÁUSULA PRIMERA: DEL PAGO DE LA MEMBRESÍA Y DÍA DE COBRO</b>", body_style))
    story.append(Paragraph(
        "El socio se compromete expresamente a cancelar el monto correspondiente de su membresía mensualmente. "
        "A partir de la presente fecha, se establece el <b>día 3 de cada mes</b> como el día límite de pago estandarizado "
        "para todos los miembros activos. Los pagos se realizarán de manera anticipada por medio de efectivo, tarjeta de "
        "crédito o débito en recepción, o bien, a través del sistema de cargo automático a tarjeta (débito recurrente) cuando "
        "esta modalidad sea habilitada por la administración del gimnasio.",
        body_style
    ))
    
    story.append(Paragraph("<b>CLÁUSULA SEGUNDA: REGLAMENTO INTERNO DE CONVIVENCIA Y SEGURIDAD</b>", body_style))
    story.append(Paragraph(
        "Para garantizar un ambiente seguro y agradable, el Socio se compromete a respetar estrictamente las normas del establecimiento:",
        body_style
    ))
    
    rules = [
        "<b>1. Prohibición de Fumar:</b> Queda terminantemente prohibido fumar o consumir cualquier tipo de vaporizador o cigarrillo electrónico dentro de todas las áreas físicas de las instalaciones.",
        "<b>2. Cuidado del Equipo:</b> El Socio deberá utilizar las máquinas, mancuernas y accesorios de manera adecuada y segura, evitando azotar o dejar caer el peso bruscamente. Todo desperfecto provocado por negligencia será responsabilidad del Socio.",
        "<b>3. Orden en Sala:</b> Es obligatorio retornar las mancuernas, barras y discos a sus respectivos racks inmediatamente al finalizar cada ejercicio.",
        "<b>4. Higiene Personal:</b> Por respeto y salud, cada Socio debe traer una toalla personal para limpiar el sudor residual en las áreas de contacto de los equipos tras su uso.",
        "<b>5. Responsabilidad:</b> Bodyflex Gym no se hace responsable por la pérdida, robo u olvido de objetos de valor o artículos personales dejados dentro de las instalaciones."
    ]
    
    for r in rules:
        story.append(Paragraph(r, bullet_style))
        
    story.append(Spacer(1, 10))
    story.append(Paragraph(
        "El incumplimiento de cualquiera de las reglas descritas anteriormente dará derecho a la administración de "
        "cancelar temporal o definitivamente la membresía del Socio sin derecho a reembolso.",
        body_style
    ))
    
    story.append(Spacer(1, 20))
    
    sig_data = [
        [
            Paragraph("_______________________________<br/><b>Firma del Socio</b><br/>CUI: " + str(usuario['cui']), body_style),
            Paragraph("_______________________________<br/><b>Por la Administración</b><br/>Bodyflex Gym", body_style)
        ]
    ]
    sig_table = Table(sig_data, colWidths=[(letter[0] - 108)/2, (letter[0] - 108)/2])
    sig_table.setStyle(TableStyle([
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('TOPPADDING', (0,0), (-1,-1), 20),
    ]))
    story.append(sig_table)
    
    doc.build(story)
    buf.seek(0)
    
    response = make_response(buf.read())
    response.headers['Content-Type'] = 'application/pdf'
    response.headers['Content-Disposition'] = f'inline; filename=contrato_{usuario["cui"]}.pdf'
    return response



if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port, debug=False)
